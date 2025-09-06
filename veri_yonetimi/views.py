from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse
from django.http import JsonResponse
from django.db.models import Q
from django.utils import timezone
from .models import AnaVeri, SutunAyarlari, Sütun, VeriDeger, UserProfile, UserLog
from .forms import AnaVeriForm, SütunForm
from django.db.models import Case, When, Value, IntegerField, CharField
from django.contrib.auth.models import User, Group

@login_required
def ana_sayfa(request):
    """
    Ana sayfa - Dashboard
    """
    # Session'a app ayarlarını set et

    
    # Aktif sütunları al
    aktif_sutunlar = Sütun.objects.filter(aktif=True).order_by('sıra')
    
    # Sütunları sıra numarasına göre sırala
    dinamik_sutunlar = aktif_sutunlar.filter(tip='dinamik').order_by('sıra')
    cihaz_sutunlar = aktif_sutunlar.filter(tip__in=['kurulacak', 'kurulan', 'arizali', 'tamamlanma']).order_by('sıra')

    # Verileri al (sadece önizleme için)
    veriler = AnaVeri.objects.all()

    # İl sorumlusu ise sadece kendi ilini göster
    if not request.user.is_superuser:
        # Kullanıcının hangi ilin sorumlusu olduğunu bul
        user_il = None
        if request.user.last_name:  # last_name'de plaka var
            user_plaka = request.user.last_name
            # Bu plakaya sahip veriyi bul
            for ana_veri in veriler:
                for deger in ana_veri.degerler.all():
                    if deger.sutun.ad == 'Plaka' and deger.deger == user_plaka:
                        user_il = ana_veri
                        break
                if user_il:
                    break

            if user_il:
                veriler = AnaVeri.objects.filter(id=user_il.id)

    # Türkiye geneli istatistikler
    turkiye_toplam_kurulacak = sum(veri.kurulacak_cihaz_sayisi for veri in AnaVeri.objects.all())
    turkiye_toplam_kurulan = sum(veri.kurulan_cihaz_sayisi for veri in AnaVeri.objects.all())
    turkiye_toplam_arizali = sum(veri.arizali_cihaz_sayisi for veri in AnaVeri.objects.all())
    turkiye_tamamlanma_yuzdesi = round((turkiye_toplam_kurulan / turkiye_toplam_kurulacak * 100), 3) if turkiye_toplam_kurulacak > 0 else 0
    
    # İl özelinde istatistikler
    il_istatistikleri = []
    for veri in veriler:
        il_istatistikleri.append({
            'veri': veri,
            'tamamlanma_yuzdesi': veri.tamamlanma_yuzdesi,
            'kalan_cihaz': veri.kalan_cihaz_sayisi,
            'durum_renk': veri.durum_renk
        })
    
    # En iyi ve en kötü performans gösteren iller
    tum_veriler = AnaVeri.objects.all()
    en_iyi_iller = sorted(tum_veriler, key=lambda x: x.tamamlanma_yuzdesi, reverse=True)[:5]
    en_kotu_iller = sorted(tum_veriler, key=lambda x: x.tamamlanma_yuzdesi)[:5]

    # İller özelinde detaylı istatistik veriler
    il_tamamlanma_verileri = []
    for veri in tum_veriler:
        if veri.kurulacak_cihaz_sayisi > 0:
            yuzde = round((veri.kurulan_cihaz_sayisi / veri.kurulacak_cihaz_sayisi * 100), 3)
            il_tamamlanma_verileri.append({
                'il': veri.il_adi,
                'yuzde': yuzde,
                'veri': veri
            })
    
    if il_tamamlanma_verileri:
        # En yüksek ve en düşük tamamlanma
        en_yuksek_tamamlanma = max(il_tamamlanma_verileri, key=lambda x: x['yuzde'])
        en_dusuk_tamamlanma = min(il_tamamlanma_verileri, key=lambda x: x['yuzde'])
        
        # Ortalama tamamlanma
        ortalama_tamamlanma = round(sum(x['yuzde'] for x in il_tamamlanma_verileri) / len(il_tamamlanma_verileri), 1)
        
        # Toplam il sayısı
        toplam_il_sayisi = len(il_tamamlanma_verileri)
        
        # Performans kategorileri
        mukemmel_il_sayisi = len([x for x in il_tamamlanma_verileri if x['yuzde'] >= 80])
        iyi_il_sayisi = len([x for x in il_tamamlanma_verileri if 60 <= x['yuzde'] < 80])
        orta_il_sayisi = len([x for x in il_tamamlanma_verileri if 40 <= x['yuzde'] < 60])
        kritik_il_sayisi = len([x for x in il_tamamlanma_verileri if x['yuzde'] < 40])
        
        # Bölgesel ortalamalar
        marmara_illeri = ['İstanbul', 'Bursa', 'Kocaeli', 'Sakarya', 'Balıkesir', 'Çanakkale', 'Edirne', 'Tekirdağ', 'Yalova', 'Kırklareli']
        ege_illeri = ['İzmir', 'Manisa', 'Aydın', 'Muğla', 'Denizli', 'Afyonkarahisar', 'Kütahya', 'Uşak']
        akdeniz_illeri = ['Antalya', 'Adana', 'Mersin', 'Hatay', 'Kahramanmaraş', 'Osmaniye', 'Isparta', 'Burdur']
        ic_anadolu_illeri = ['Ankara', 'Konya', 'Kayseri', 'Sivas', 'Yozgat', 'Kırıkkale', 'Aksaray', 'Niğde', 'Nevşehir', 'Kırşehir', 'Çankırı', 'Karaman']
        
        marmara_ortalamasi = round(sum(x['yuzde'] for x in il_tamamlanma_verileri if x['il'] in marmara_illeri) / max(len([x for x in il_tamamlanma_verileri if x['il'] in marmara_illeri]), 1), 1)
        ege_ortalamasi = round(sum(x['yuzde'] for x in il_tamamlanma_verileri if x['il'] in ege_illeri) / max(len([x for x in il_tamamlanma_verileri if x['il'] in ege_illeri]), 1), 1)
        akdeniz_ortalamasi = round(sum(x['yuzde'] for x in il_tamamlanma_verileri if x['il'] in akdeniz_illeri) / max(len([x for x in il_tamamlanma_verileri if x['il'] in akdeniz_illeri]), 1), 1)
        ic_anadolu_ortalamasi = round(sum(x['yuzde'] for x in il_tamamlanma_verileri if x['il'] in ic_anadolu_illeri) / max(len([x for x in il_tamamlanma_verileri if x['il'] in ic_anadolu_illeri]), 1), 1)
    else:
        en_yuksek_tamamlanma = {'yuzde': 0, 'il': '-'}
        en_dusuk_tamamlanma = {'yuzde': 0, 'il': '-'}
        ortalama_tamamlanma = 0
        toplam_il_sayisi = 0
        mukemmel_il_sayisi = 0
        iyi_il_sayisi = 0
        orta_il_sayisi = 0
        kritik_il_sayisi = 0
        marmara_ortalamasi = 0
        ege_ortalamasi = 0
        akdeniz_ortalamasi = 0
        ic_anadolu_ortalamasi = 0

    # Sağ sidebar için ek veriler
    user_count = User.objects.count()

    context = {
        'veriler': veriler,
        'aktif_sutunlar': aktif_sutunlar,
        'dinamik_sutunlar': dinamik_sutunlar,
        'cihaz_sutunlar': cihaz_sutunlar,
        'user_count': user_count,
        'turkiye_toplam_kurulacak': turkiye_toplam_kurulacak,
        'turkiye_toplam_kurulan': turkiye_toplam_kurulan,
        'turkiye_toplam_arizali': turkiye_toplam_arizali,
        'turkiye_tamamlanma_yuzdesi': turkiye_tamamlanma_yuzdesi,
        'il_istatistikleri': il_istatistikleri,
        'en_iyi_iller': en_iyi_iller,
        'en_kotu_iller': en_kotu_iller,
        'en_yuksek_tamamlanma': en_yuksek_tamamlanma,
        'en_dusuk_tamamlanma': en_dusuk_tamamlanma,
        'ortalama_tamamlanma': ortalama_tamamlanma,
        'toplam_il_sayisi': toplam_il_sayisi,
        'mukemmel_il_sayisi': mukemmel_il_sayisi,
        'iyi_il_sayisi': iyi_il_sayisi,
        'orta_il_sayisi': orta_il_sayisi,
        'kritik_il_sayisi': kritik_il_sayisi,
        'marmara_ortalamasi': marmara_ortalamasi,
        'ege_ortalamasi': ege_ortalamasi,
        'akdeniz_ortalamasi': akdeniz_ortalamasi,
        'ic_anadolu_ortalamasi': ic_anadolu_ortalamasi,


    }
    return render(request, 'veri_yonetimi/ana_sayfa.html', context)

@login_required
def veri_listesi(request):
    """
    Ana veri tablosunu listele
    """
    # Session'a app ayarlarını set et

    
    # Filtreleme parametreleri
    search_query = request.GET.get('search', '')
    sort_by = request.GET.get('sort', 'olusturulma_tarihi')
    sort_order = request.GET.get('order', 'desc')
    
    # Sütun bazında filtreleme
    sutun_filtreleri = {}
    aktif_sutunlar = Sütun.objects.filter(aktif=True).order_by('sıra')
    
    # Sütunları sıra numarasına göre sırala
    dinamik_sutunlar = aktif_sutunlar.filter(tip='dinamik').order_by('sıra')
    cihaz_sutunlar = aktif_sutunlar.filter(tip__in=['kurulacak', 'kurulan', 'arizali', 'tamamlanma']).order_by('sıra')
    
    for sutun in dinamik_sutunlar:
        filter_value = request.GET.get(f'filter_{sutun.id}', '')
        if filter_value:
            sutun_filtreleri[sutun.id] = filter_value
    
    # Verileri al - İl bazlı filtreleme
    veriler = AnaVeri.objects.all()
    
    # Superuser değilse sadece sorumlu olduğu iller verilerini göster
    if not request.user.is_superuser:
        try:
            user_profile = request.user.profile
            sorumlu_iller = user_profile.get_sorumlu_iller_list()
            if sorumlu_iller:
                # Kullanıcının sorumlu olduğu illere göre filtrele
                veriler = veriler.filter(il_adi__in=sorumlu_iller)
            # Eğer sorumlu_iller boşsa: tüm illeri görebilir
        except:
            # UserProfile yoksa hiçbir veri gösterme
            veriler = AnaVeri.objects.none()
    
    # Genel arama filtresi
    if search_query:
        veriler = veriler.filter(
            Q(degerler__deger__icontains=search_query) |
            Q(id__icontains=search_query)
        ).distinct()
    
    # Sütun bazında filtreleme
    for sutun_id, filter_value in sutun_filtreleri.items():
        veriler = veriler.filter(
            degerler__sutun_id=sutun_id,
            degerler__deger__icontains=filter_value
        ).distinct()
    
    # Sıralama
    if sort_by == 'olusturulma_tarihi':
        if sort_order == 'desc':
            veriler = veriler.order_by('-olusturulma_tarihi')
        else:
            veriler = veriler.order_by('olusturulma_tarihi')
    elif sort_by == 'id':
        if sort_order == 'desc':
            veriler = veriler.order_by('-id')
        else:
            veriler = veriler.order_by('id')
    elif sort_by.startswith('sutun_'):
        # Sütun bazında sıralama
        sutun_id = sort_by.replace('sutun_', '')
        
        # Sütun değerlerini al ve sırala
        if sort_order == 'desc':
            # Azalan sıralama için önce boş değerler, sonra dolu değerler
            veriler = veriler.annotate(
                sutun_deger=Case(
                    When(degerler__sutun_id=sutun_id, then='degerler__deger'),
                    default=Value(''),
                    output_field=CharField(),
                )
            ).order_by('sutun_deger', '-olusturulma_tarihi')
        else:
            # Artan sıralama için önce dolu değerler, sonra boş değerler
            veriler = veriler.annotate(
                sutun_deger=Case(
                    When(degerler__sutun_id=sutun_id, then='degerler__deger'),
                    default=Value(''),
                    output_field=CharField(),
                )
            ).order_by('sutun_deger', 'olusturulma_tarihi')
    else:
        # Varsayılan sıralama: Plaka sırasına göre
        veriler = veriler.annotate(
            plaka_deger=Case(
                When(degerler__sutun__ad='Plaka', then='degerler__deger'),
                default=Value(''),
                output_field=CharField(),
            )
        ).order_by('plaka_deger')
    
    # Eğer hiç sütun yoksa varsayılan sütunları oluştur
    if not aktif_sutunlar.exists():
        varsayilan_sutunlar = [
            {'ad': 'Plaka', 'sıra': 1},
            {'ad': 'İl Adı', 'sıra': 2},
            {'ad': 'Kurulacak Cihaz Sayısı', 'sıra': 3},
            {'ad': 'Kurulan Cihaz Sayısı', 'sıra': 4},
            {'ad': 'Arızalı Cihaz Sayısı', 'sıra': 5},
        ]
        for sutun_data in varsayilan_sutunlar:
            Sütun.objects.create(**sutun_data)
        aktif_sutunlar = Sütun.objects.filter(aktif=True).order_by('sıra')
    
    # Türkiye geneli istatistikler
    turkiye_toplam_kurulacak = sum(veri.kurulacak_cihaz_sayisi for veri in AnaVeri.objects.all())
    turkiye_toplam_kurulan = sum(veri.kurulan_cihaz_sayisi for veri in AnaVeri.objects.all())
    turkiye_toplam_arizali = sum(veri.arizali_cihaz_sayisi for veri in AnaVeri.objects.all())
    turkiye_tamamlanma_yuzdesi = round((turkiye_toplam_kurulan / turkiye_toplam_kurulacak * 100), 3) if turkiye_toplam_kurulacak > 0 else 0
    
    # İller özelinde istatistik veriler
    from django.db.models import Avg, Count, Q
    
    # Tüm illerin tamamlanma yüzdeleri
    il_tamamlanma_verileri = []
    for veri in AnaVeri.objects.all():
        if veri.kurulacak_cihaz_sayisi > 0:
            yuzde = round((veri.kurulan_cihaz_sayisi / veri.kurulacak_cihaz_sayisi * 100), 3)
            il_tamamlanma_verileri.append({
                'il': veri.il_adi,
                'yuzde': yuzde,
                'veri': veri
            })
    
    if il_tamamlanma_verileri:
        # En yüksek ve en düşük tamamlanma
        en_yuksek_tamamlanma = max(il_tamamlanma_verileri, key=lambda x: x['yuzde'])
        en_dusuk_tamamlanma = min(il_tamamlanma_verileri, key=lambda x: x['yuzde'])
        
        # Ortalama tamamlanma
        ortalama_tamamlanma = round(sum(x['yuzde'] for x in il_tamamlanma_verileri) / len(il_tamamlanma_verileri), 1)
        
        # Toplam il sayısı
        toplam_il_sayisi = len(il_tamamlanma_verileri)
        
        # Performans kategorileri
        mukemmel_il_sayisi = len([x for x in il_tamamlanma_verileri if x['yuzde'] >= 80])
        iyi_il_sayisi = len([x for x in il_tamamlanma_verileri if 60 <= x['yuzde'] < 80])
        orta_il_sayisi = len([x for x in il_tamamlanma_verileri if 40 <= x['yuzde'] < 60])
        kritik_il_sayisi = len([x for x in il_tamamlanma_verileri if x['yuzde'] < 40])
        
        # Bölgesel ortalamalar (basit hesaplama)
        marmara_illeri = ['İstanbul', 'Bursa', 'Kocaeli', 'Sakarya', 'Balıkesir', 'Çanakkale', 'Edirne', 'Tekirdağ', 'Yalova', 'Kırklareli']
        ege_illeri = ['İzmir', 'Manisa', 'Aydın', 'Muğla', 'Denizli', 'Afyonkarahisar', 'Kütahya', 'Uşak']
        akdeniz_illeri = ['Antalya', 'Adana', 'Mersin', 'Hatay', 'Kahramanmaraş', 'Osmaniye', 'Isparta', 'Burdur']
        ic_anadolu_illeri = ['Ankara', 'Konya', 'Kayseri', 'Sivas', 'Yozgat', 'Kırıkkale', 'Aksaray', 'Niğde', 'Nevşehir', 'Kırşehir', 'Çankırı', 'Karaman']
        
        marmara_ortalamasi = round(sum(x['yuzde'] for x in il_tamamlanma_verileri if x['il'] in marmara_illeri) / max(len([x for x in il_tamamlanma_verileri if x['il'] in marmara_illeri]), 1), 1)
        ege_ortalamasi = round(sum(x['yuzde'] for x in il_tamamlanma_verileri if x['il'] in ege_illeri) / max(len([x for x in il_tamamlanma_verileri if x['il'] in ege_illeri]), 1), 1)
        akdeniz_ortalamasi = round(sum(x['yuzde'] for x in il_tamamlanma_verileri if x['il'] in akdeniz_illeri) / max(len([x for x in il_tamamlanma_verileri if x['il'] in akdeniz_illeri]), 1), 1)
        ic_anadolu_ortalamasi = round(sum(x['yuzde'] for x in il_tamamlanma_verileri if x['il'] in ic_anadolu_illeri) / max(len([x for x in il_tamamlanma_verileri if x['il'] in ic_anadolu_illeri]), 1), 1)
    else:
        en_yuksek_tamamlanma = {'yuzde': 0, 'il': '-'}
        en_dusuk_tamamlanma = {'yuzde': 0, 'il': '-'}
        ortalama_tamamlanma = 0
        toplam_il_sayisi = 0
        mukemmel_il_sayisi = 0
        iyi_il_sayisi = 0
        orta_il_sayisi = 0
        kritik_il_sayisi = 0
        marmara_ortalamasi = 0
        ege_ortalamasi = 0
        akdeniz_ortalamasi = 0
        ic_anadolu_ortalamasi = 0
    
    # Sağ sidebar için ek veriler
    user_count = User.objects.count()
    sutun_count = Sütun.objects.filter(aktif=True).count()
    toplam_veri_count = AnaVeri.objects.count()
    bugun_eklenen = AnaVeri.objects.filter(olusturulma_tarihi__date=timezone.now().date()).count()
    
    context = {
        'veriler': veriler,
        'sutunlar': Sütun.get_menu_columns('veri_listesi'),
        'aktif_sutunlar': aktif_sutunlar,
        'dinamik_sutunlar': dinamik_sutunlar,
        'cihaz_sutunlar': cihaz_sutunlar,
        'search_query': search_query,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'turkiye_toplam_kurulacak': turkiye_toplam_kurulacak,
        'turkiye_toplam_kurulan': turkiye_toplam_kurulan,
        'turkiye_toplam_arizali': turkiye_toplam_arizali,
        'turkiye_tamamlanma_yuzdesi': turkiye_tamamlanma_yuzdesi,
        'en_yuksek_tamamlanma': en_yuksek_tamamlanma,
        'en_dusuk_tamamlanma': en_dusuk_tamamlanma,
        'ortalama_tamamlanma': ortalama_tamamlanma,
        'toplam_il_sayisi': toplam_il_sayisi,
        'mukemmel_il_sayisi': mukemmel_il_sayisi,
        'iyi_il_sayisi': iyi_il_sayisi,
        'orta_il_sayisi': orta_il_sayisi,
        'kritik_il_sayisi': kritik_il_sayisi,
        'marmara_ortalamasi': marmara_ortalamasi,
        'ege_ortalamasi': ege_ortalamasi,
        'akdeniz_ortalamasi': akdeniz_ortalamasi,
        'ic_anadolu_ortalamasi': ic_anadolu_ortalamasi,
        'user_count': user_count,
        'sutun_count': sutun_count,
        'toplam_veri_count': toplam_veri_count,
        'bugun_eklenen': bugun_eklenen,


    }
    return render(request, 'veri_yonetimi/veri_listesi.html', context)

@login_required
def veri_ekle(request):
    """
    Yeni veri ekle
    """
    from django.contrib import messages
    
    # Session'a app ayarlarını set et

    
    # Login kontrolü
    if not request.user.is_authenticated:
        messages.error(request, 'Veri eklemek için giriş yapmalısınız.')
        return redirect('veri_yonetimi:user_login')
    
    if request.method == 'POST':
        
        form = AnaVeriForm(request.POST, user=request.user)
        
        if form.is_valid():
            try:
                # Form'u doğrudan kaydet (AnaVeri ve VeriDeger'lar birlikte)
                ana_veri = form.save()
                
                # VeriDeger'ları kontrol et
                for deger in ana_veri.degerler.all():
                    pass
                
                # Log işlemi kaydet
                log_user_activity(
                    user=request.user,
                    islem_yapan=request.user,
                    islem_tipi='veri_eklendi',
                    aciklama=f'{request.user.username} kullanıcısı yeni veri ekledi (ID: {ana_veri.id})',
                    request=request,
                    yeni_deger={
                        'ana_veri_id': ana_veri.id,
                        'il_adi': ana_veri.il_adi,
                        'kurulacak_cihaz_sayisi': ana_veri.kurulacak_cihaz_sayisi,
                        'kurulan_cihaz_sayisi': ana_veri.kurulan_cihaz_sayisi,
                        'arizali_cihaz_sayisi': ana_veri.arizali_cihaz_sayisi
                    }
                )
                
                # Detaylı başarı mesajı oluştur
                success_message = f'📊 Veri Başarıyla Eklendi!|İl: {ana_veri.il_adi}|Kurulacak: {ana_veri.kurulacak_cihaz_sayisi:,} cihaz|Kurulan: {ana_veri.kurulan_cihaz_sayisi:,} cihaz|Tamamlanma: %{ana_veri.tamamlanma_yuzdesi}|Tarih: {timezone.now().strftime("%d.%m.%Y %H:%M")}'
                messages.success(request, success_message)
                return redirect('veri_yonetimi:veri_listesi')
            except Exception as e:
                import traceback
                traceback.print_exc()
                messages.error(request, f'Veri eklenirken hata oluştu: {str(e)}')
                # Hata detaylarını log'la
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'Veri ekleme hatası: {str(e)}')
                logger.error(f'Form data: {request.POST}')
        else:
            for field_name, errors in form.errors.items():
                print(f"  - {field_name}: {errors}")
    else:
        form = AnaVeriForm(user=request.user)
    
    # Aktif sütunları al
    aktif_sutunlar = Sütun.objects.filter(aktif=True).order_by('sıra')
    
    # Sağ sidebar için ek veriler
    user_count = User.objects.count()
    
    context = {
        'form': form,
        'veri': None,  # Yeni veri için None
        'aktif_sutunlar': aktif_sutunlar,
        'is_edit': False,
        'user_count': user_count,
    }
    return render(request, 'veri_yonetimi/veri_formu.html', context)

@login_required
def veri_guncelle(request, pk):
    """
    Veri güncelle
    """
    from django.contrib import messages
    
    # Session'a app ayarlarını set et

    
    # Login kontrolü
    if not request.user.is_authenticated:
        messages.error(request, 'Veri güncellemek için giriş yapmalısınız.')
        return redirect('veri_yonetimi:user_login')
    
    ana_veri = get_object_or_404(AnaVeri, pk=pk)
    
    # İl sorumluluğu kontrolü (superuser değilse)
    if not request.user.is_superuser:
        try:
            user_profile = request.user.profile
            if not user_profile.is_responsible_for_il(ana_veri.il_adi):
                sorumlu_iller = user_profile.get_sorumlu_iller_display()
                messages.error(request, f'Sadece sorumlu olduğunuz iller ({sorumlu_iller}) verilerini güncelleyebilirsiniz.')
                return redirect('veri_yonetimi:veri_listesi')
        except:
            messages.error(request, 'Profil bilgileriniz eksik. Yönetici ile iletişime geçin.')
            return redirect('veri_yonetimi:veri_listesi')
    
    if request.method == 'POST':
        form = AnaVeriForm(request.POST, user=request.user, instance=ana_veri)
        if form.is_valid():
            try:
                # Eski değerleri kaydet
                eski_degerler = {
                    'il_adi': ana_veri.il_adi,
                    'kurulacak_cihaz_sayisi': ana_veri.kurulacak_cihaz_sayisi,
                    'kurulan_cihaz_sayisi': ana_veri.kurulan_cihaz_sayisi,
                    'arizali_cihaz_sayisi': ana_veri.arizali_cihaz_sayisi
                }
                
                # Form'u doğrudan kaydet (AnaVeri ve VeriDeger'lar birlikte)
                ana_veri = form.save()
                
                # Yeni değerler
                yeni_degerler = {
                    'il_adi': ana_veri.il_adi,
                    'kurulacak_cihaz_sayisi': ana_veri.kurulacak_cihaz_sayisi,
                    'kurulan_cihaz_sayisi': ana_veri.kurulan_cihaz_sayisi,
                    'arizali_cihaz_sayisi': ana_veri.arizali_cihaz_sayisi
                }
                
                # Log işlemi kaydet
                log_user_activity(
                    user=request.user,
                    islem_yapan=request.user,
                    islem_tipi='veri_guncellendi',
                    aciklama=f'{request.user.username} kullanıcısı veri güncelledi (ID: {ana_veri.id})',
                    request=request,
                    eski_deger=eski_degerler,
                    yeni_deger=yeni_degerler
                )
                
                # Detaylı başarı mesajı oluştur
                success_message = f'📊 Veri Başarıyla Güncellendi!|İl: {ana_veri.il_adi}|Kurulacak: {ana_veri.kurulacak_cihaz_sayisi:,} cihaz|Kurulan: {ana_veri.kurulan_cihaz_sayisi:,} cihaz|Tamamlanma: %{ana_veri.tamamlanma_yuzdesi}|Tarih: {timezone.now().strftime("%d.%m.%Y %H:%M")}'
                messages.success(request, success_message)
                return redirect('veri_yonetimi:veri_listesi')
            except Exception as e:
                messages.error(request, f'Veri güncellenirken hata oluştu: {str(e)}')
                # Hata detaylarını log'la
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'Veri güncelleme hatası: {str(e)}')
                logger.error(f'Form data: {request.POST}')
    else:
        # Mevcut veriyi form'a yükle
        initial_data = {}
        form = AnaVeriForm(user=request.user, instance=ana_veri)
        form.set_initial_from_instance()  # Instance'dan değerleri yükle
    
    # Aktif sütunları al
    aktif_sutunlar = Sütun.objects.filter(aktif=True).order_by('sıra')
    
    # Form field değerlerini template için hazırla
    field_values = {}
    for field_name, field in form.fields.items():
        field_values[field_name] = form.initial.get(field_name, '')
    
    # JSON olarak serialize et
    import json
    field_values_json = json.dumps(field_values)
    
    # Sağ sidebar için ek veriler
    user_count = User.objects.count()
    
    context = {
        'form': form,
        'veri': ana_veri,  # Template'de 'veri' olarak kullanılıyor
        'ana_veri': ana_veri,
        'aktif_sutunlar': aktif_sutunlar,
        'field_values': field_values,
        'field_values_json': field_values_json,
        'is_edit': True,
        'user_count': user_count,
    }
    return render(request, 'veri_yonetimi/veri_formu.html', context)

@login_required
def veri_sil(request, pk):
    """
    Veri silme onay sayfası
    """
    # Session'a app ayarlarını set et

    
    # Login kontrolü
    if not request.user.is_authenticated:
        messages.error(request, 'Veri silmek için giriş yapmalısınız.')
        return redirect('veri_yonetimi:user_login')
    
    ana_veri = get_object_or_404(AnaVeri, pk=pk)
    
    # İl sorumluluğu kontrolü (superuser değilse)
    if not request.user.is_superuser:
        try:
            user_profile = request.user.profile
            if not user_profile.is_responsible_for_il(ana_veri.il_adi):
                sorumlu_iller = user_profile.get_sorumlu_iller_display()
                messages.error(request, f'Sadece sorumlu olduğunuz iller ({sorumlu_iller}) verilerini silebilirsiniz.')
                return redirect('veri_yonetimi:veri_listesi')
        except:
            messages.error(request, 'Profil bilgileriniz eksik. Yönetici ile iletişime geçin.')
            return redirect('veri_yonetimi:veri_listesi')
    
    if request.method == 'POST':
        try:
            # Silinen veri bilgilerini kaydet
            silinen_veri = {
                'ana_veri_id': ana_veri.id,
                'il_adi': ana_veri.il_adi,
                'kurulacak_cihaz_sayisi': ana_veri.kurulacak_cihaz_sayisi,
                'kurulan_cihaz_sayisi': ana_veri.kurulan_cihaz_sayisi,
                'arizali_cihaz_sayisi': ana_veri.arizali_cihaz_sayisi
            }
            
            # Log işlemi kaydet (silmeden önce)
            log_user_activity(
                user=request.user,
                islem_yapan=request.user,
                islem_tipi='veri_silindi',
                aciklama=f'{request.user.username} kullanıcısı veri sildi (ID: {ana_veri.id}, İl: {ana_veri.il_adi})',
                request=request,
                eski_deger=silinen_veri
            )
            
            # Veriyi sil
            ana_veri.delete()
            messages.success(request, 'Veri başarıyla silindi!')
            return redirect('veri_yonetimi:veri_listesi')
        except Exception as e:
            messages.error(request, f'Veri silinirken hata oluştu: {str(e)}')
            return redirect('veri_yonetimi:veri_listesi')
    
    # Aktif sütunları al
    aktif_sutunlar = Sütun.objects.filter(aktif=True).order_by('sıra')
    
    context = {
        'ana_veri': ana_veri,
        'aktif_sutunlar': aktif_sutunlar,
    }
    return render(request, 'veri_yonetimi/veri_sil.html', context)

@login_required
def veri_sil_onay(request, pk):
    """
    AJAX ile veri silme onayı
    """
    if request.method == 'POST':
        veri = get_object_or_404(AnaVeri, pk=pk)
        veri.delete()
        return JsonResponse({'success': True, 'message': 'Veri başarıyla silindi!'})
    
    return JsonResponse({'success': False, 'message': 'Geçersiz istek!'})

# Sütun yönetimi view'ları
@login_required
def sutun_listesi(request):
    """
    Sütun listesi
    """
    # Session'a app ayarlarını set et


    
    sutunlar = Sütun.objects.all().order_by('sıra')
    
    # Uygulama ayarlarını veritabanından al
    from .models import AppSettings
    app_settings = AppSettings.get_settings()
    site_title = app_settings['app_name']
    
    # Sağ sidebar için ek veriler
    user_count = User.objects.count()
    aktif_sutunlar = Sütun.objects.filter(aktif=True).order_by('sıra')
    
    context = {
        'sutunlar': sutunlar,
        'site_title': site_title,
        'user_count': user_count,
        'aktif_sutunlar': aktif_sutunlar,
    }
    return render(request, 'veri_yonetimi/sutun_listesi.html', context)

@login_required
def sutun_ekle(request):
    """
    Yeni sütun ekle
    """
    
    if request.method == 'POST':
        form = SütunForm(request.POST)
        if form.is_valid():
            sutun = form.save()
            
            # Log işlemi kaydet
            log_user_activity(
                user=request.user,
                islem_yapan=request.user,
                islem_tipi='sutun_eklendi',
                aciklama=f'{request.user.username} kullanıcısı yeni sütun ekledi: {sutun.ad}',
                request=request,
                yeni_deger={
                    'sutun_id': sutun.id,
                    'ad': sutun.ad,
                    'tip': sutun.tip,
                    'menu_tipi': sutun.menu_tipi,
                    'sıra': sutun.sıra
                }
            )
            
            # Detaylı başarı mesajı oluştur
            success_message = f'➕ Sütun Başarıyla Eklendi!|Ad: {sutun.ad}|Tip: {sutun.tip}|Menü Tipi: {sutun.menu_tipi}|Sıra: {sutun.sıra}|Tarih: {timezone.now().strftime("%d.%m.%Y %H:%M")}'
            messages.success(request, success_message)
            return redirect('veri_yonetimi:sutun_listesi')
        else:
            messages.error(request, 'Form hataları var. Lütfen kontrol edin.')
    else:
        form = SütunForm()
    
    # Sağ sidebar için ek veriler
    user_count = User.objects.count()
    aktif_sutunlar = Sütun.objects.filter(aktif=True).order_by('sıra')
    
    context = {
        'form': form,
        'is_edit': False,
        'user_count': user_count,
        'aktif_sutunlar': aktif_sutunlar,
    }
    return render(request, 'veri_yonetimi/sutun_formu.html', context)

@login_required
def sutun_guncelle(request, pk):
    """
    Sütun güncelle
    """
    # Session'a app ayarlarını set et

    
    sutun = get_object_or_404(Sütun, pk=pk)
    
    if request.method == 'POST':
        # Eski değerleri kaydet
        eski_degerler = {
            'ad': sutun.ad,
            'tip': sutun.tip,
            'menu_tipi': sutun.menu_tipi,
            'sıra': sutun.sıra,
            'aktif': sutun.aktif
        }
        
        form = SütunForm(request.POST, instance=sutun)
        if form.is_valid():
            sutun = form.save()
            
            # Yeni değerler
            yeni_degerler = {
                'ad': sutun.ad,
                'tip': sutun.tip,
                'menu_tipi': sutun.menu_tipi,
                'sıra': sutun.sıra,
                'aktif': sutun.aktif
            }
            
            # Log işlemi kaydet
            log_user_activity(
                user=request.user,
                islem_yapan=request.user,
                islem_tipi='sutun_guncellendi',
                aciklama=f'{request.user.username} kullanıcısı sütun güncelledi: {sutun.ad}',
                request=request,
                eski_deger=eski_degerler,
                yeni_deger=yeni_degerler
            )
            
            # Detaylı başarı mesajı oluştur
            success_message = f'✏️ Sütun Başarıyla Güncellendi!|Ad: {sutun.ad}|Tip: {sutun.tip}|Menü Tipi: {sutun.menu_tipi}|Sıra: {sutun.sıra}|Aktif: {"Evet" if sutun.aktif else "Hayır"}|Tarih: {timezone.now().strftime("%d.%m.%Y %H:%M")}'
            messages.success(request, success_message)
            return redirect('veri_yonetimi:sutun_listesi')
        else:
            messages.error(request, 'Form hataları var. Lütfen kontrol edin.')
    else:
        form = SütunForm(instance=sutun)
    
    # Sağ sidebar için ek veriler
    user_count = User.objects.count()
    aktif_sutunlar = Sütun.objects.filter(aktif=True).order_by('sıra')
    
    context = {
        'form': form,
        'sutun': sutun,
        'is_edit': True,
        'user_count': user_count,
        'aktif_sutunlar': aktif_sutunlar,
    }
    return render(request, 'veri_yonetimi/sutun_formu.html', context)

@login_required
def sutun_sil(request, pk):
    """
    Sütun sil
    """
    # Session'a app ayarlarını set et

    
    try:
        sutun = get_object_or_404(Sütun, pk=pk)
        
        if request.method == 'POST':
            
            try:
                # Önce VeriDeger kayıtlarını sil (eğer varsa)
                from .models import VeriDeger
                veri_degerleri = VeriDeger.objects.filter(sutun=sutun)
                veri_sayisi = veri_degerleri.count()
                
                if veri_sayisi > 0:
                    veri_degerleri.delete()
                
                # Silinen sütun bilgilerini kaydet
                silinen_sutun = {
                    'sutun_id': sutun.id,
                    'ad': sutun.ad,
                    'tip': sutun.tip,
                    'menu_tipi': sutun.menu_tipi,
                    'sıra': sutun.sıra
                }
                
                # Log işlemi kaydet (silmeden önce)
                log_user_activity(
                    user=request.user,
                    islem_yapan=request.user,
                    islem_tipi='sutun_silindi',
                    aciklama=f'{request.user.username} kullanıcısı sütun sildi: {sutun.ad}',
                    request=request,
                    eski_deger=silinen_sutun
                )
                
                # Sütunu sil
                sutun.delete()
                messages.success(request, 'Sütun başarıyla silindi!')
                return redirect('veri_yonetimi:sutun_listesi')
                
            except Exception as e:
                import traceback
                traceback.print_exc()
                messages.error(request, f'Sütun silinirken hata oluştu: {str(e)}')
                return redirect('veri_yonetimi:sutun_listesi')
        
        
        # Sağ sidebar için ek veriler
        user_count = User.objects.count()
        aktif_sutunlar = Sütun.objects.filter(aktif=True).order_by('sıra')
        
        context = {
            'sutun': sutun,
            'user_count': user_count,
            'aktif_sutunlar': aktif_sutunlar,
        }
        return render(request, 'veri_yonetimi/sutun_sil.html', context)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        messages.error(request, f'Sütun bulunamadı: {str(e)}')
        return redirect('veri_yonetimi:sutun_listesi')

@login_required
def update_site_title(request):
    """
    Site başlığını güncelle
    """
    if request.method == 'POST':
        site_title = request.POST.get('site_title', '').strip()
        
        if site_title:
            # Site başlığını session'a kaydet (gerçek uygulamada veritabanına kaydedilir)
            request.session['site_title'] = site_title
            messages.success(request, f'Site başlığı başarıyla güncellendi: {site_title}')
        else:
            messages.error(request, 'Site başlığı boş olamaz!')
    
    return redirect('veri_yonetimi:sutun_listesi')

@login_required
def update_app_settings(request):
    """
    Uygulama ayarlarını güncelle
    """
    if request.method == 'POST':
        app_name = request.POST.get('app_name', '').strip()
        app_description = request.POST.get('app_description', '').strip()
        app_logo_file = request.FILES.get('app_logo')
        
        if app_name:
            # Uygulama ayarlarını veritabanına kaydet
            from .models import AppSettings
            
            try:
                AppSettings.update_settings(
                    app_name=app_name,
                    app_description=app_description,
                    app_logo=app_logo_file
                )
                messages.success(request, f'Uygulama ayarları başarıyla güncellendi: {app_name}')
            except Exception as e:
                messages.error(request, f'Ayarlar kaydedilirken hata oluştu: {str(e)}')
        else:
            messages.error(request, 'Uygulama adı boş olamaz!')
    
    return redirect('veri_yonetimi:sutun_listesi')

@login_required
def kullanici_listesi(request):
    """
    Kullanıcı listesini göster (sadece admin)
    """
    # Session'a app ayarlarını set et

    
    # Sadece süper kullanıcılar kullanıcı listesini görebilir
    if not request.user.is_superuser:
        from django.contrib import messages
        messages.error(request, 'Sadece admin kullanıcılar kullanıcı listesini görebilir.')
        return redirect('veri_yonetimi:veri_listesi')
    
    # Tüm kullanıcıları al
    kullanicilar = User.objects.all().order_by('username')
    
    # İl sorumlusu grubunu al
    try:
        il_sorumlusu_group = Group.objects.get(name='İl Sorumlusu')
    except Group.DoesNotExist:
        il_sorumlusu_group = None
    
    # Sağ sidebar için ek veriler
    son_veriler = AnaVeri.objects.order_by('-olusturulma_tarihi')[:5]
    user_count = User.objects.count()
    aktif_sutunlar = Sütun.objects.filter(aktif=True).order_by('sıra')
    
    # İstatistikler
    admin_count = User.objects.filter(is_superuser=True).count()
    active_count = User.objects.filter(is_active=True).count()
    il_sorumlusu_count = User.objects.filter(is_staff=True, is_superuser=False).count()
    
    context = {
        'kullanicilar': kullanicilar,
        'il_sorumlusu_group': il_sorumlusu_group,
        'son_veriler': son_veriler,
        'user_count': user_count,
        'aktif_sutunlar': aktif_sutunlar,
        'admin_count': admin_count,
        'active_count': active_count,
        'il_sorumlusu_count': il_sorumlusu_count,
    }
    return render(request, 'veri_yonetimi/kullanici_listesi.html', context)

@login_required 
def kullanici_ekle(request):
    """
    Yeni kullanıcı ekle
    """
    # Sadece yetkili kullanıcılar erişebilir
    if not request.user.is_superuser:
        messages.error(request, 'Bu sayfaya erişim yetkiniz yok.')
        return redirect('veri_yonetimi:ana_sayfa')
    
    if request.method == 'POST':
        # Form verilerini al
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        tc_kimlik = request.POST.get('tc_kimlik')
        sorumlu_iller = request.POST.getlist('sorumlu_iller')
        user_role = request.POST.get('user_role', 'viewer')
        
        # Şifre kontrolü
        if password1 != password2:
            messages.error(request, 'Şifreler eşleşmiyor.')
            return redirect('veri_yonetimi:kullanici_ekle')
        
        # Kullanıcı adı kontrolü
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Bu kullanıcı adı zaten kullanılıyor.')
            return redirect('veri_yonetimi:kullanici_ekle')
        
        # TC kimlik kontrolü
        if tc_kimlik and UserProfile.objects.filter(tc_kimlik=tc_kimlik).exists():
            messages.error(request, 'Bu TC kimlik numarası zaten kullanılıyor.')
            return redirect('veri_yonetimi:kullanici_ekle')
        
        try:
            # Kullanıcıyı oluştur
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password1,
                first_name=first_name,
                last_name=last_name
            )
            
            # Role göre yetkilendirme ayarla
            if user_role == 'super_user':
                user.is_superuser = True
                user.is_staff = True
            elif user_role == 'province_admin':
                user.is_staff = True
                user.is_superuser = False
            elif user_role == 'province_manager':
                user.is_staff = True
                user.is_superuser = False
            else:  # viewer
                user.is_staff = False
                user.is_superuser = False
                
            user.save()
            
            # UserProfile oluştur
            profile = UserProfile.objects.create(
                user=user, 
                tc_kimlik=tc_kimlik if tc_kimlik else None,
                role=user_role
            )
            # Sorumlu illeri set et
            if sorumlu_iller:
                profile.set_sorumlu_iller(sorumlu_iller)
                profile.save()
            
            # Log işlemi kaydet
            log_user_activity(
                user=user,
                islem_yapan=request.user,
                islem_tipi='kullanici_olusturuldu',
                aciklama=f'"{username}" kullanıcısı oluşturuldu. TC: {tc_kimlik or "Belirtilmemiş"}, Email: {email or "Belirtilmemiş"}, Rol: {profile.get_role_display()}',
                request=request,
                yeni_deger={
                    'username': username,
                    'email': email,
                    'first_name': first_name,
                    'last_name': last_name,
                    'is_superuser': user.is_superuser,
                    'is_staff': user.is_staff,
                    'tc_kimlik': tc_kimlik,
                    'role': user_role
                }
            )
            
            messages.success(request, f'"{username}" kullanıcısı başarıyla oluşturuldu!')
            return redirect('veri_yonetimi:kullanici_listesi')
            
        except Exception as e:
            messages.error(request, f'Kullanıcı oluşturulurken hata oluştu: {str(e)}')
            return redirect('veri_yonetimi:kullanici_ekle')
    
    # Türkiye illeri
    turkiye_illeri = [
        'Adana', 'Adıyaman', 'Afyonkarahisar', 'Ağrı', 'Amasya', 'Ankara', 'Antalya', 'Artvin', 
        'Aydın', 'Balıkesir', 'Bilecik', 'Bingöl', 'Bitlis', 'Bolu', 'Burdur', 'Bursa', 
        'Çanakkale', 'Çankırı', 'Çorum', 'Denizli', 'Diyarbakır', 'Edirne', 'Elazığ', 'Erzincan', 
        'Erzurum', 'Eskişehir', 'Gaziantep', 'Giresun', 'Gümüşhane', 'Hakkari', 'Hatay', 'Isparta', 
        'Mersin', 'İstanbul', 'İzmir', 'Kars', 'Kastamonu', 'Kayseri', 'Kırklareli', 'Kırşehir', 
        'Kocaeli', 'Konya', 'Kütahya', 'Malatya', 'Manisa', 'Kahramanmaraş', 'Mardin', 'Muğla', 
        'Muş', 'Nevşehir', 'Niğde', 'Ordu', 'Rize', 'Sakarya', 'Samsun', 'Siirt', 'Sinop', 
        'Sivas', 'Tekirdağ', 'Tokat', 'Trabzon', 'Tunceli', 'Şanlıurfa', 'Uşak', 'Van', 
        'Yozgat', 'Zonguldak', 'Aksaray', 'Bayburt', 'Karaman', 'Kırıkkale', 'Batman', 'Şırnak', 
        'Bartın', 'Ardahan', 'Iğdır', 'Yalova', 'Karabük', 'Kilis', 'Osmaniye', 'Düzce'
    ]
    
    context = {
        'turkiye_illeri': sorted(turkiye_illeri),
    }
    return render(request, 'veri_yonetimi/kullanici_formu.html', context)

@login_required
def kullanici_guncelle(request, pk):
    """
    Kullanıcı güncelle
    """
    # Session'a app ayarlarını set et

    
    # Sadece süper kullanıcılar kullanıcı güncelleyebilir
    if not request.user.is_superuser:
        messages.error(request, 'Sadece admin kullanıcılar kullanıcı güncelleyebilir.')
        return redirect('veri_yonetimi:kullanici_listesi')
    
    kullanici = get_object_or_404(User, pk=pk)
    
    if request.method == 'POST':
        # Form verilerini al
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        tc_kimlik = request.POST.get('tc_kimlik')
        is_superuser = request.POST.get('is_superuser') == 'on'
        is_staff = request.POST.get('is_staff') == 'on'
        is_active = request.POST.get('is_active') == 'on'
        
        # TC kimlik benzersizlik kontrolü
        current_tc = getattr(kullanici.profile, 'tc_kimlik', None) if hasattr(kullanici, 'profile') else None
        if tc_kimlik and tc_kimlik != current_tc:
            if UserProfile.objects.filter(tc_kimlik=tc_kimlik).exclude(user=kullanici).exists():
                messages.error(request, 'Bu TC kimlik numarası başka bir kullanıcı tarafından kullanılıyor.')
                return redirect('veri_yonetimi:kullanici_guncelle', pk=pk)
        
        try:
            # Eski değerleri kaydet
            eski_degerler = {
                'email': kullanici.email,
                'first_name': kullanici.first_name,
                'last_name': kullanici.last_name,
                'is_superuser': kullanici.is_superuser,
                'is_staff': kullanici.is_staff,
                'is_active': kullanici.is_active,
                'tc_kimlik': getattr(kullanici.profile, 'tc_kimlik', None) if hasattr(kullanici, 'profile') else None
            }
            
            # Kullanıcıyı güncelle
            kullanici.email = email
            kullanici.first_name = first_name
            kullanici.last_name = last_name
            kullanici.is_superuser = is_superuser
            kullanici.is_staff = is_staff
            kullanici.is_active = is_active
            kullanici.save()
            
            # TC kimlik numarasını güncelle (UserProfile ile)
            if hasattr(kullanici, 'profile'):
                kullanici.profile.tc_kimlik = tc_kimlik
                kullanici.profile.save()
            elif tc_kimlik:
                UserProfile.objects.create(user=kullanici, tc_kimlik=tc_kimlik)
            
            # Yeni değerler
            yeni_degerler = {
                'email': email,
                'first_name': first_name,
                'last_name': last_name,
                'is_superuser': is_superuser,
                'is_staff': is_staff,
                'is_active': is_active,
                'tc_kimlik': tc_kimlik
            }
            
            # Log işlemi kaydet
            log_user_activity(
                user=kullanici,
                islem_yapan=request.user,
                islem_tipi='kullanici_guncellendi',
                aciklama=f'"{kullanici.username}" kullanıcısının bilgileri güncellendi.',
                request=request,
                eski_deger=eski_degerler,
                yeni_deger=yeni_degerler
            )
            
            # Detaylı başarı mesajı oluştur
            role_text = 'Yönetici' if is_superuser else 'İl Sorumlusu' if is_staff else 'Kullanıcı'
            success_message = f'👤 Kullanıcı Başarıyla Güncellendi!|Kullanıcı Adı: {kullanici.username}|Ad Soyad: {(first_name + " " + last_name).strip() or "Belirtilmemiş"}|E-posta: {email or "Belirtilmemiş"}|Rol: {role_text}|Durum: {"Aktif" if is_active else "Pasif"}|Tarih: {timezone.now().strftime("%d.%m.%Y %H:%M")}'
            messages.success(request, success_message)
            return redirect('veri_yonetimi:kullanici_listesi')
            
        except Exception as e:
            messages.error(request, f'Kullanıcı güncellenirken hata oluştu: {str(e)}')
            return redirect('veri_yonetimi:kullanici_guncelle', pk=pk)
    
    # Sağ sidebar için ek veriler
    user_count = User.objects.count()
    aktif_sutunlar = Sütun.objects.filter(aktif=True).order_by('sıra')
    
    context = {
        'kullanici': kullanici,
        'user_count': user_count,
        'aktif_sutunlar': aktif_sutunlar,
    }
    return render(request, 'veri_yonetimi/kullanici_formu.html', context)

@login_required
def generate_fake_tc(request):
    """
    Tüm kullanıcılar için fake TC kimlik numaraları üret
    """
    # Sadece süper kullanıcılar fake TC üretebilir
    if not request.user.is_superuser:
        messages.error(request, 'Sadece admin kullanıcılar fake TC kimlik numaraları üretebilir.')
        return redirect('veri_yonetimi:kullanici_listesi')
    
    if request.method == 'POST':
        try:
            from .models import UserProfile
            import random
            
            # Mevcut kullanıcıları al
            users = User.objects.all()
            created_count = 0
            updated_count = 0
            
            for user in users:
                # Fake TC kimlik numarası oluştur
                fake_tc = generate_valid_tc()
                
                # UserProfile oluştur veya güncelle
                if hasattr(user, 'profile'):
                    user.profile.tc_kimlik = fake_tc
                    user.profile.save()
                    updated_count += 1
                else:
                    UserProfile.objects.create(user=user, tc_kimlik=fake_tc)
                    created_count += 1
            
            messages.success(request, f'🎲 {created_count} yeni, {updated_count} güncellenmiş fake TC kimlik numarası oluşturuldu!')
            
        except Exception as e:
            messages.error(request, f'Fake TC kimlik numaraları oluşturulurken hata oluştu: {str(e)}')
    
    return redirect('veri_yonetimi:kullanici_listesi')

def generate_valid_tc():
    """
    Geçerli fake TC kimlik numarası oluştur
    """
    from .models import UserProfile
    import random
    
    while True:
        # İlk 9 haneyi rastgele oluştur
        first_nine = ''.join([str(random.randint(0, 9)) for _ in range(9)])
        
        # 10. hane (1. kontrol hanesi)
        sum_odd = sum(int(first_nine[i]) for i in range(0, 9, 2))
        sum_even = sum(int(first_nine[i]) for i in range(1, 8, 2))
        
        digit_10 = (sum_odd * 7 - sum_even) % 10
        
        # 11. hane (2. kontrol hanesi)
        first_ten = first_nine + str(digit_10)
        sum_all = sum(int(first_ten[i]) for i in range(10))
        
        digit_11 = sum_all % 10
        
        # Tam TC kimlik numarası
        tc = first_nine + str(digit_10) + str(digit_11)
        
        # Benzersizlik kontrolü
        if not UserProfile.objects.filter(tc_kimlik=tc).exists():
            return tc



@login_required
def cihaz_turleri_duzenle(request, pk):
    """
    Cihaz türü düzenle
    """
    # Session'a app ayarlarını set et

    
    # Cihaz türleri verileri
    cihaz_turleri_data = [
        {
            'id': 1,
            'ad': 'Sürücü Analiz Kamerası',
            'aciklama': 'Sürücü davranışlarını analiz eden kamera sistemi',
            'kategori': 'Güvenlik',
            'durum': 'Aktif',
            'kurulum_sayisi': 1250,
            'hedef_sayisi': 2000,
            'tamamlanma_yuzdesi': 62.5,
            'icon': '📹',
            'renk': 'blue'
        },
        {
            'id': 2,
            'ad': 'ATS Araç Takip Sistemi',
            'aciklama': 'Araç konum ve durum takip sistemi',
            'kategori': 'Takip',
            'durum': 'Aktif',
            'kurulum_sayisi': 980,
            'hedef_sayisi': 1500,
            'tamamlanma_yuzdesi': 65.3,
            'icon': '🚗',
            'renk': 'green'
        },
        {
            'id': 3,
            'ad': 'GPS Konum Takip',
            'aciklama': 'Hassas konum belirleme ve rota takibi',
            'kategori': 'Navigasyon',
            'durum': 'Aktif',
            'kurulum_sayisi': 2100,
            'hedef_sayisi': 2500,
            'tamamlanma_yuzdesi': 84.0,
            'icon': '📍',
            'renk': 'purple'
        },
        {
            'id': 4,
            'ad': 'Yakıt Takip Sistemi',
            'aciklama': 'Yakıt tüketimi ve maliyet takibi',
            'kategori': 'Maliyet',
            'durum': 'Aktif',
            'kurulum_sayisi': 750,
            'hedef_sayisi': 1200,
            'tamamlanma_yuzdesi': 62.5,
            'icon': '⛽',
            'renk': 'orange'
        },
        {
            'id': 5,
            'ad': 'Motor Performans Monitörü',
            'aciklama': 'Motor sağlığı ve performans takibi',
            'kategori': 'Teknik',
            'durum': 'Aktif',
            'kurulum_sayisi': 680,
            'hedef_sayisi': 1000,
            'tamamlanma_yuzdesi': 68.0,
            'icon': '⚙️',
            'renk': 'red'
        },
        {
            'id': 6,
            'ad': 'Hız ve Mesafe Sensörü',
            'aciklama': 'Hız limiti ve güvenlik uyarıları',
            'kategori': 'Güvenlik',
            'durum': 'Aktif',
            'kurulum_sayisi': 890,
            'hedef_sayisi': 1300,
            'tamamlanma_yuzdesi': 68.5,
            'icon': '🏃',
            'renk': 'indigo'
        }
    ]
    
    # PK'ya göre cihaz türünü bul
    cihaz_turu = None
    for cihaz in cihaz_turleri_data:
        if cihaz['id'] == pk:
            cihaz_turu = cihaz
            break
    
    if not cihaz_turu:
        messages.error(request, 'Cihaz türü bulunamadı!')
        return redirect('veri_yonetimi:cihaz_turleri')
    
    if request.method == 'POST':
        # Form verilerini al
        cihaz_turu['ad'] = request.POST.get('ad', cihaz_turu['ad'])
        cihaz_turu['aciklama'] = request.POST.get('aciklama', cihaz_turu['aciklama'])
        cihaz_turu['kategori'] = request.POST.get('kategori', cihaz_turu['kategori'])
        cihaz_turu['durum'] = request.POST.get('durum', cihaz_turu['durum'])
        cihaz_turu['kurulum_sayisi'] = int(request.POST.get('kurulum_sayisi', cihaz_turu['kurulum_sayisi']))
        cihaz_turu['hedef_sayisi'] = int(request.POST.get('hedef_sayisi', cihaz_turu['hedef_sayisi']))
        cihaz_turu['icon'] = request.POST.get('icon', cihaz_turu['icon'])
        
        # Tamamlanma yüzdesini hesapla
        if cihaz_turu['hedef_sayisi'] > 0:
            cihaz_turu['tamamlanma_yuzdesi'] = round((cihaz_turu['kurulum_sayisi'] / cihaz_turu['hedef_sayisi']) * 100, 1)
        
        messages.success(request, 'Cihaz türü başarıyla güncellendi!')
        return redirect('veri_yonetimi:cihaz_turleri')
    
    context = {
        'cihaz_turu': cihaz_turu,
        'kategoriler': ['Güvenlik', 'Takip', 'Navigasyon', 'Maliyet', 'Teknik'],
        'durumlar': ['Aktif', 'Pasif'],
        'icons': ['📹', '🚗', '📍', '⛽', '⚙️', '🏃', '🔧', '📡', '💻', '📱'],

    }
    
    return render(request, 'veri_yonetimi/cihaz_turleri_duzenle.html', context)

@login_required
def cihaz_turleri_sil(request, pk):
    """
    Cihaz türü sil
    """
    # Session'a app ayarlarını set et

    
    # Cihaz türleri verileri
    cihaz_turleri_data = [
        {
            'id': 1,
            'ad': 'Sürücü Analiz Kamerası',
            'aciklama': 'Sürücü davranışlarını analiz eden kamera sistemi',
            'kategori': 'Güvenlik',
            'durum': 'Aktif',
            'kurulum_sayisi': 1250,
            'hedef_sayisi': 2000,
            'tamamlanma_yuzdesi': 62.5,
            'icon': '📹',
            'renk': 'blue'
        },
        {
            'id': 2,
            'ad': 'ATS Araç Takip Sistemi',
            'aciklama': 'Araç konum ve durum takip sistemi',
            'kategori': 'Takip',
            'durum': 'Aktif',
            'kurulum_sayisi': 980,
            'hedef_sayisi': 1500,
            'tamamlanma_yuzdesi': 65.3,
            'icon': '🚗',
            'renk': 'green'
        },
        {
            'id': 3,
            'ad': 'GPS Konum Takip',
            'aciklama': 'Hassas konum belirleme ve rota takibi',
            'kategori': 'Navigasyon',
            'durum': 'Aktif',
            'kurulum_sayisi': 2100,
            'hedef_sayisi': 2500,
            'tamamlanma_yuzdesi': 84.0,
            'icon': '📍',
            'renk': 'purple'
        },
        {
            'id': 4,
            'ad': 'Yakıt Takip Sistemi',
            'aciklama': 'Yakıt tüketimi ve maliyet takibi',
            'kategori': 'Maliyet',
            'durum': 'Aktif',
            'kurulum_sayisi': 750,
            'hedef_sayisi': 1200,
            'tamamlanma_yuzdesi': 62.5,
            'icon': '⛽',
            'renk': 'orange'
        },
        {
            'id': 5,
            'ad': 'Motor Performans Monitörü',
            'aciklama': 'Motor sağlığı ve performans takibi',
            'kategori': 'Teknik',
            'durum': 'Aktif',
            'kurulum_sayisi': 680,
            'hedef_sayisi': 1000,
            'tamamlanma_yuzdesi': 68.0,
            'icon': '⚙️',
            'renk': 'red'
        },
        {
            'id': 6,
            'ad': 'Hız ve Mesafe Sensörü',
            'aciklama': 'Hız limiti ve güvenlik uyarıları',
            'kategori': 'Güvenlik',
            'durum': 'Aktif',
            'kurulum_sayisi': 890,
            'hedef_sayisi': 1300,
            'tamamlanma_yuzdesi': 68.5,
            'icon': '🏃',
            'renk': 'indigo'
        }
    ]
    
    # PK'ya göre cihaz türünü bul
    cihaz_turu = None
    for cihaz in cihaz_turleri_data:
        if cihaz['id'] == pk:
            cihaz_turu = cihaz
            break
    
    if not cihaz_turu:
        messages.error(request, 'Cihaz türü bulunamadı!')
        return redirect('veri_yonetimi:cihaz_turleri')
    
    if request.method == 'POST':
        # Cihaz türünü sil (gerçek uygulamada veritabanından silinir)
        messages.success(request, f'{cihaz_turu["ad"]} cihaz türü başarıyla silindi!')
        return redirect('veri_yonetimi:cihaz_turleri')
    
    context = {
        'cihaz_turu': cihaz_turu,

    }
    
    return render(request, 'veri_yonetimi/cihaz_turleri_sil.html', context)

@login_required
def toggle_user_status(request, pk):
    """
    Kullanıcı durumunu aktif/pasif yap
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Bu işlem için yetkiniz yok'})
    
    try:
        user = get_object_or_404(User, pk=pk)
        # Kendini pasif yapamaz
        if user == request.user:
            return JsonResponse({'success': False, 'error': 'Kendinizi pasif yapamazsınız'})
        
        # Eski durumu kaydet
        eski_durum = user.is_active
        
        # Durumu değiştir
        user.is_active = not user.is_active
        user.save()
        
        # Log işlemi kaydet
        log_user_activity(
            user=user,
            islem_yapan=request.user,
            islem_tipi='durum_degistirildi',
            aciklama=f'{user.username} kullanıcısı {"aktif" if user.is_active else "pasif"} yapıldı',
            request=request,
            eski_deger={'is_active': eski_durum},
            yeni_deger={'is_active': user.is_active}
        )
        
        status_text = 'aktif' if user.is_active else 'pasif'
        return JsonResponse({
            'success': True, 
            'message': f'{user.username} kullanıcısı {status_text} yapıldı',
            'new_status': user.is_active
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def change_user_role(request, pk):
    """
    Kullanıcı yetkisini değiştir
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Bu işlem için yetkiniz yok'})
    
    try:
        import json
        data = json.loads(request.body)
        new_role = data.get('role')
        
        user = get_object_or_404(User, pk=pk)
        # Kendinin yetkisini değiştiremez
        if user == request.user:
            return JsonResponse({'success': False, 'error': 'Kendi yetkinizi değiştiremezsiniz'})
        
        # Eski yetkileri kaydet
        eski_yetkiler = {
            'is_superuser': user.is_superuser,
            'is_staff': user.is_staff
        }
        
        # Yetkiyi değiştir
        if new_role == 'normal':
            user.is_superuser = False
            user.is_staff = False
        elif new_role == 'staff':
            user.is_superuser = False
            user.is_staff = True
        elif new_role == 'superuser':
            user.is_superuser = True
            user.is_staff = True
        else:
            return JsonResponse({'success': False, 'error': 'Geçersiz yetki seviyesi'})
        
        user.save()
        
        # Log işlemi kaydet
        role_names = {
            'normal': 'Normal Kullanıcı',
            'staff': 'İl Sorumlusu',
            'superuser': 'Admin'
        }
        
        log_user_activity(
            user=user,
            islem_yapan=request.user,
            islem_tipi='yetki_degistirildi',
            aciklama=f'{user.username} kullanıcısının yetkisi {role_names[new_role]} olarak değiştirildi',
            request=request,
            eski_deger=eski_yetkiler,
            yeni_deger={'is_superuser': user.is_superuser, 'is_staff': user.is_staff}
        )
        
        return JsonResponse({
            'success': True,
            'message': f'{user.username} kullanıcısının yetkisi {role_names[new_role]} olarak değiştirildi',
            'new_role': new_role
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def get_client_ip(request):
    """İstemci IP adresini al"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def log_user_activity(user, islem_yapan, islem_tipi, aciklama, request=None, eski_deger=None, yeni_deger=None):
    """Kullanıcı işlemlerini logla"""
    try:
        ip_adresi = None
        user_agent = None
        
        if request:
            ip_adresi = get_client_ip(request)
            user_agent = request.META.get('HTTP_USER_AGENT', '')
        
        UserLog.objects.create(
            user=user,
            islem_yapan=islem_yapan,
            islem_tipi=islem_tipi,
            aciklama=aciklama,
            ip_adresi=ip_adresi,
            user_agent=user_agent,
            eski_deger=eski_deger,
            yeni_deger=yeni_deger
        )
    except Exception as e:
        # Log hatası sistem loglarına yazdırılabilir
        print(f"Log kaydı oluşturulamadı: {str(e)}")

@login_required
def kullanici_loglari(request):
    """
    Kullanıcı işlem loglarını listele
    """
    from django.utils import timezone
    
    # Session'a app ayarlarını set et

    
    # Sadece süper kullanıcılar logları görebilir
    if not request.user.is_superuser:
        messages.error(request, 'Sadece admin kullanıcılar işlem loglarını görebilir.')
        return redirect('veri_yonetimi:kullanici_listesi')
    
    # Filtreleme parametreleri
    search_query = request.GET.get('search', '')
    user_filter = request.GET.get('user_filter', '')
    islem_filter = request.GET.get('islem_filter', '')
    tarih_baslangic = request.GET.get('tarih_baslangic', '')
    tarih_bitis = request.GET.get('tarih_bitis', '')
    
    # Logları al
    loglar = UserLog.objects.all().select_related('user', 'islem_yapan')
    
    # Filtreleme
    if search_query:
        loglar = loglar.filter(
            Q(aciklama__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(islem_yapan__username__icontains=search_query)
        )
    
    if user_filter:
        loglar = loglar.filter(user__username__icontains=user_filter)
    
    if islem_filter:
        loglar = loglar.filter(islem_tipi=islem_filter)
    
    if tarih_baslangic:
        from datetime import datetime
        try:
            baslangic_tarihi = datetime.strptime(tarih_baslangic, '%Y-%m-%d')
            # Timezone aware yapma
            baslangic_tarihi = timezone.make_aware(baslangic_tarihi)
            loglar = loglar.filter(tarih__gte=baslangic_tarihi)
        except:
            pass
    
    if tarih_bitis:
        from datetime import datetime
        try:
            bitis_tarihi = datetime.strptime(tarih_bitis, '%Y-%m-%d')
            # Günün sonuna kadar dahil et
            from datetime import timedelta
            bitis_tarihi = bitis_tarihi + timedelta(days=1)
            # Timezone aware yapma
            bitis_tarihi = timezone.make_aware(bitis_tarihi)
            loglar = loglar.filter(tarih__lt=bitis_tarihi)
        except:
            pass
    
    # İstatistikler (pagination'dan ÖNCE)
    toplam_log_sayisi = UserLog.objects.count()
    bugun_log_sayisi = UserLog.objects.filter(tarih__date=timezone.now().date()).count()
    aktif_kullanici_sayisi = User.objects.filter(is_active=True).count()
    
    # En çok işlem yapan kullanıcılar
    from django.db.models import Count
    en_aktif_kullanicilar = UserLog.objects.values('islem_yapan__username').annotate(
        islem_sayisi=Count('id')
    ).order_by('-islem_sayisi')[:5]
    
    # İşlem tipi dağılımı
    islem_tipi_dagilimi = UserLog.objects.values('islem_tipi').annotate(
        sayi=Count('id')
    ).order_by('-sayi')[:10]
    
    # Kullanıcı listesini al
    all_users = User.objects.all().order_by('username')
    
    # Sayfalama (pagination) - son adım
    from django.core.paginator import Paginator
    paginator = Paginator(loglar, 20)  # Sayfa başına 20 log
    sayfa_no = request.GET.get('page')
    loglar = paginator.get_page(sayfa_no)
    
    context = {
        'loglar': loglar,
        'islem_tipleri': UserLog.ISLEM_TIPLERI,
        'kullanicilar': all_users,
        'search_query': search_query,
        'user_filter': user_filter,
        'islem_filter': islem_filter,
        'tarih_baslangic': tarih_baslangic,
        'tarih_bitis': tarih_bitis,
        'toplam_log_sayisi': toplam_log_sayisi,
        'bugun_log_sayisi': bugun_log_sayisi,
        'aktif_kullanici_sayisi': aktif_kullanici_sayisi,
        'en_aktif_kullanicilar': en_aktif_kullanicilar,
        'islem_tipi_dagilimi': islem_tipi_dagilimi,


    }
    return render(request, 'veri_yonetimi/kullanici_loglari.html', context)

@login_required
def kullanici_loglari_excel(request):
    """
    Kullanıcı loglarını Excel formatında indir
    """
    # Sadece süper kullanıcılar logları indirebilir
    if not request.user.is_superuser:
        messages.error(request, 'Sadece admin kullanıcılar işlem loglarını indirebilir.')
        return redirect('veri_yonetimi:kullanici_listesi')
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from datetime import datetime
    
    # Filtreleme parametreleri (aynı filtreleri uygula)
    search_query = request.GET.get('search', '')
    user_filter = request.GET.get('user_filter', '')
    islem_filter = request.GET.get('islem_filter', '')
    tarih_baslangic = request.GET.get('tarih_baslangic', '')
    tarih_bitis = request.GET.get('tarih_bitis', '')
    
    # Logları al
    loglar = UserLog.objects.all().select_related('user', 'islem_yapan')
    
    # Aynı filtreleri uygula
    if search_query:
        loglar = loglar.filter(
            Q(aciklama__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(islem_yapan__username__icontains=search_query)
        )
    
    if user_filter:
        loglar = loglar.filter(user__username__icontains=user_filter)
    
    if islem_filter:
        loglar = loglar.filter(islem_tipi=islem_filter)
    
    if tarih_baslangic:
        try:
            baslangic_tarihi = datetime.strptime(tarih_baslangic, '%Y-%m-%d')
            # Timezone aware yapma
            from django.utils import timezone as tz
            baslangic_tarihi = tz.make_aware(baslangic_tarihi)
            loglar = loglar.filter(tarih__gte=baslangic_tarihi)
        except:
            pass
    
    if tarih_bitis:
        try:
            bitis_tarihi = datetime.strptime(tarih_bitis, '%Y-%m-%d')
            from datetime import timedelta
            bitis_tarihi = bitis_tarihi + timedelta(days=1)
            # Timezone aware yapma
            from django.utils import timezone as tz
            bitis_tarihi = tz.make_aware(bitis_tarihi)
            loglar = loglar.filter(tarih__lt=bitis_tarihi)
        except:
            pass
    
    # Excel dosyası oluştur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kullanıcı İşlem Logları"
    
    # Başlık satırları
    headers = [
        'Sıra No',
        'Kullanıcı',
        'İşlemi Yapan',
        'İşlem Tipi',
        'Açıklama',
        'IP Adresi',
        'Tarih',
        'Saat'
    ]
    
    # Başlık stilini ayarla
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Veri satırlarını ekle
    for row, log in enumerate(loglar, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=log.user.username)
        ws.cell(row=row, column=3, value=log.islem_yapan.username if log.islem_yapan else 'Sistem')
        ws.cell(row=row, column=4, value=log.get_islem_tipi_display())
        ws.cell(row=row, column=5, value=log.aciklama)
        ws.cell(row=row, column=6, value=log.ip_adresi or '-')
        ws.cell(row=row, column=7, value=log.tarih.strftime('%d.%m.%Y'))
        ws.cell(row=row, column=8, value=log.tarih.strftime('%H:%M:%S'))
    
    # Sütun genişliklerini ayarla
    column_widths = [8, 15, 15, 20, 40, 15, 12, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # HTTP response oluştur
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f'kullanici_loglari_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    
    # Log kaydı oluştur
    log_user_activity(
        user=request.user,
        islem_yapan=request.user,
        islem_tipi='excel_indirildi',
        aciklama=f'Kullanıcı işlem logları Excel formatında indirildi. Toplam {loglar.count()} kayıt.',
        request=request
    )
    
    return response


# ===========================
# KULLANICI YÖNETİMİ VİEW'LARI
# ===========================

@login_required
def kullanici_listesi(request):
    """
    Kullanıcı listesi
    """
    # Sadece superuser erişebilir
    if not request.user.is_superuser:
        messages.error(request, 'Bu sayfaya erişim yetkiniz yok.')
        return redirect('veri_yonetimi:ana_sayfa')
    
    # Session'a app ayarlarını set et

    
    # Kullanıcıları al
    kullanicilar = User.objects.all().order_by('-date_joined')
    
    # Türkiye illeri listesi
    turkiye_illeri = [
        'Adana', 'Adıyaman', 'Afyonkarahisar', 'Ağrı', 'Amasya', 'Ankara', 'Antalya', 'Artvin', 
        'Aydın', 'Balıkesir', 'Bilecik', 'Bingöl', 'Bitlis', 'Bolu', 'Burdur', 'Bursa', 
        'Çanakkale', 'Çankırı', 'Çorum', 'Denizli', 'Diyarbakır', 'Edirne', 'Elazığ', 'Erzincan', 
        'Erzurum', 'Eskişehir', 'Gaziantep', 'Giresun', 'Gümüşhane', 'Hakkari', 'Hatay', 'Isparta', 
        'Mersin', 'İstanbul', 'İzmir', 'Kars', 'Kastamonu', 'Kayseri', 'Kırklareli', 'Kırşehir', 
        'Kocaeli', 'Konya', 'Kütahya', 'Malatya', 'Manisa', 'Kahramanmaraş', 'Mardin', 'Muğla', 
        'Muş', 'Nevşehir', 'Niğde', 'Ordu', 'Rize', 'Sakarya', 'Samsun', 'Siirt', 'Sinop', 
        'Sivas', 'Tekirdağ', 'Tokat', 'Trabzon', 'Tunceli', 'Şanlıurfa', 'Uşak', 'Van', 'Yozgat', 
        'Zonguldak', 'Aksaray', 'Bayburt', 'Karaman', 'Kırıkkale', 'Batman', 'Şırnak', 'Bartın', 
        'Ardahan', 'Iğdır', 'Yalova', 'Karabük', 'Kilis', 'Osmaniye', 'Düzce'
    ]
    
    # İstatistikler
    admin_count = kullanicilar.filter(is_superuser=True).count()
    staff_count = kullanicilar.filter(is_staff=True, is_superuser=False).count()
    user_count = kullanicilar.filter(is_staff=False, is_superuser=False).count()
    aktif_count = kullanicilar.filter(is_active=True).count()
    
    context = {
        'kullanicilar': kullanicilar,
        'turkiye_illeri': turkiye_illeri,
        'admin_count': admin_count,
        'staff_count': staff_count, 
        'user_count': user_count,
        'aktif_count': aktif_count,
    }
    return render(request, 'veri_yonetimi/kullanici_listesi.html', context)


@login_required
def kullanici_ekle(request):
    """
    Yeni kullanıcı ekle
    """
    # Sadece superuser erişebilir
    if not request.user.is_superuser:
        messages.error(request, 'Bu sayfaya erişim yetkiniz yok.')
        return redirect('veri_yonetimi:ana_sayfa')
    
    if request.method == 'POST':
        # Form verilerini al
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password = request.POST.get('password')
        tc_kimlik = request.POST.get('tc_kimlik')
        sorumlu_iller = request.POST.getlist('sorumlu_iller')
        role = request.POST.get('role', 'user')
        
        try:
            # Kullanıcı oluştur
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name
            )
            
            # Role ayarla
            if role == 'admin':
                user.is_superuser = True
                user.is_staff = True
            elif role == 'staff':
                user.is_staff = True
                user.is_superuser = False
            else:
                user.is_staff = False
                user.is_superuser = False
                
            user.save()
            
            # UserProfile oluştur veya güncelle
            profile, created = UserProfile.objects.get_or_create(user=user)
            profile.tc_kimlik = tc_kimlik
            profile.set_sorumlu_iller(sorumlu_iller)
            profile.save()
            
            # Log kaydet
            log_user_activity(
                user=user,
                islem_yapan=request.user,
                islem_tipi='kullanici_eklendi',
                aciklama=f'Yeni kullanıcı eklendi: {username} ({role})',
                request=request
            )
            
            messages.success(request, f'Kullanıcı "{username}" başarıyla oluşturuldu.')
            return redirect('veri_yonetimi:kullanici_listesi')
            
        except Exception as e:
            messages.error(request, f'Kullanıcı oluşturulurken hata: {str(e)}')
    
    # Türkiye illeri listesi
    turkiye_illeri = [
        'Adana', 'Adıyaman', 'Afyonkarahisar', 'Ağrı', 'Amasya', 'Ankara', 'Antalya', 'Artvin', 
        'Aydın', 'Balıkesir', 'Bilecik', 'Bingöl', 'Bitlis', 'Bolu', 'Burdur', 'Bursa', 
        'Çanakkale', 'Çankırı', 'Çorum', 'Denizli', 'Diyarbakır', 'Edirne', 'Elazığ', 'Erzincan', 
        'Erzurum', 'Eskişehir', 'Gaziantep', 'Giresun', 'Gümüşhane', 'Hakkari', 'Hatay', 'Isparta', 
        'Mersin', 'İstanbul', 'İzmir', 'Kars', 'Kastamonu', 'Kayseri', 'Kırklareli', 'Kırşehir', 
        'Kocaeli', 'Konya', 'Kütahya', 'Malatya', 'Manisa', 'Kahramanmaraş', 'Mardin', 'Muğla', 
        'Muş', 'Nevşehir', 'Niğde', 'Ordu', 'Rize', 'Sakarya', 'Samsun', 'Siirt', 'Sinop', 
        'Sivas', 'Tekirdağ', 'Tokat', 'Trabzon', 'Tunceli', 'Şanlıurfa', 'Uşak', 'Van', 'Yozgat', 
        'Zonguldak', 'Aksaray', 'Bayburt', 'Karaman', 'Kırıkkale', 'Batman', 'Şırnak', 'Bartın', 
        'Ardahan', 'Iğdır', 'Yalova', 'Karabük', 'Kilis', 'Osmaniye', 'Düzce'
    ]
    
    context = {
        'turkiye_illeri': turkiye_illeri,
    }
    return render(request, 'veri_yonetimi/kullanici_formu.html', context)


@login_required 
def kullanici_guncelle(request, pk):
    """
    Kullanıcı güncelle
    """
    # Sadece superuser erişebilir
    if not request.user.is_superuser:
        messages.error(request, 'Bu sayfaya erişim yetkiniz yok.')
        return redirect('veri_yonetimi:ana_sayfa')
    
    user = get_object_or_404(User, pk=pk)
    profile, created = UserProfile.objects.get_or_create(user=user)
    
    if request.method == 'POST':
        # Form verilerini al
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        tc_kimlik = request.POST.get('tc_kimlik')
        sorumlu_iller = request.POST.getlist('sorumlu_iller')
        user_role = request.POST.get('user_role', 'viewer')
        
        try:
            # Kullanıcı adı kontrolü (kendisi hariç)
            if User.objects.filter(username=username).exclude(pk=user.pk).exists():
                messages.error(request, 'Bu kullanıcı adı zaten kullanılıyor.')
                return redirect('veri_yonetimi:kullanici_guncelle', pk=pk)
            
            # TC kimlik kontrolü (kendisi hariç)
            if tc_kimlik and UserProfile.objects.filter(tc_kimlik=tc_kimlik).exclude(user=user).exists():
                messages.error(request, 'Bu TC kimlik numarası zaten kullanılıyor.')
                return redirect('veri_yonetimi:kullanici_guncelle', pk=pk)
            
            # Eski değerleri kaydet
            eski_degerler = {
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'is_superuser': user.is_superuser,
                'is_staff': user.is_staff,
                'tc_kimlik': getattr(user.profile, 'tc_kimlik', None) if hasattr(user, 'profile') else None,
                'role': getattr(user.profile, 'role', 'viewer') if hasattr(user, 'profile') else 'viewer'
            }
            
            # Kullanıcı bilgilerini güncelle
            user.username = username
            user.email = email
            user.first_name = first_name
            user.last_name = last_name
            
            # Role göre yetkilendirme ayarla
            if user_role == 'super_user':
                user.is_superuser = True
                user.is_staff = True
            elif user_role == 'province_admin':
                user.is_staff = True
                user.is_superuser = False
            elif user_role == 'province_manager':
                user.is_staff = True
                user.is_superuser = False
            else:  # viewer
                user.is_staff = False
                user.is_superuser = False
                
            user.save()
            
            # Profile güncelle
            profile.tc_kimlik = tc_kimlik
            profile.role = user_role
            profile.set_sorumlu_iller(sorumlu_iller)
            profile.save()
            
            # Yeni değerler
            yeni_degerler = {
                'username': username,
                'email': email,
                'first_name': first_name,
                'last_name': last_name,
                'is_superuser': user.is_superuser,
                'is_staff': user.is_staff,
                'tc_kimlik': tc_kimlik,
                'role': user_role
            }
            
            # Log kaydet
            log_user_activity(
                user=user,
                islem_yapan=request.user,
                islem_tipi='kullanici_guncellendi',
                aciklama=f'Kullanıcı güncellendi: {username} ({profile.get_role_display()})',
                request=request,
                eski_deger=eski_degerler,
                yeni_deger=yeni_degerler
            )
            
            messages.success(request, f'Kullanıcı "{username}" başarıyla güncellendi.')
            return redirect('veri_yonetimi:kullanici_listesi')
            
        except Exception as e:
            messages.error(request, f'Kullanıcı güncellenirken hata: {str(e)}')
    
    # Türkiye illeri listesi
    turkiye_illeri = [
        'Adana', 'Adıyaman', 'Afyonkarahisar', 'Ağrı', 'Amasya', 'Ankara', 'Antalya', 'Artvin', 
        'Aydın', 'Balıkesir', 'Bilecik', 'Bingöl', 'Bitlis', 'Bolu', 'Burdur', 'Bursa', 
        'Çanakkale', 'Çankırı', 'Çorum', 'Denizli', 'Diyarbakır', 'Edirne', 'Elazığ', 'Erzincan', 
        'Erzurum', 'Eskişehir', 'Gaziantep', 'Giresun', 'Gümüşhane', 'Hakkari', 'Hatay', 'Isparta', 
        'Mersin', 'İstanbul', 'İzmir', 'Kars', 'Kastamonu', 'Kayseri', 'Kırklareli', 'Kırşehir', 
        'Kocaeli', 'Konya', 'Kütahya', 'Malatya', 'Manisa', 'Kahramanmaraş', 'Mardin', 'Muğla', 
        'Muş', 'Nevşehir', 'Niğde', 'Ordu', 'Rize', 'Sakarya', 'Samsun', 'Siirt', 'Sinop', 
        'Sivas', 'Tekirdağ', 'Tokat', 'Trabzon', 'Tunceli', 'Şanlıurfa', 'Uşak', 'Van', 
        'Yozgat', 'Zonguldak', 'Aksaray', 'Bayburt', 'Karaman', 'Kırıkkale', 'Batman', 'Şırnak', 
        'Bartın', 'Ardahan', 'Iğdır', 'Yalova', 'Karabük', 'Kilis', 'Osmaniye', 'Düzce'
    ]
    
    context = {
        'user_obj': user,
        'profile': profile,
        'turkiye_illeri': sorted(turkiye_illeri),
        'is_edit': True,
    }
    return render(request, 'veri_yonetimi/kullanici_formu.html', context)


@login_required
def toggle_user_status(request, pk):
    """
    Kullanıcı durumunu aktif/pasif yap
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Yetki hatası'})
    
    if request.method == 'POST':
        try:
            user = get_object_or_404(User, pk=pk)
            
            # Kendisini pasif yapmasını engelle
            if user == request.user:
                return JsonResponse({'success': False, 'error': 'Kendinizi pasif yapamazsınız'})
            
            import json
            data = json.loads(request.body)
            user.is_active = data.get('is_active', True)
            user.save()
            
            # Log kaydet
            durum = 'aktif' if user.is_active else 'pasif'
            log_user_activity(
                user=user,
                islem_yapan=request.user,
                islem_tipi='kullanici_durum_degisti',
                aciklama=f'Kullanıcı durumu {durum} yapıldı: {user.username}',
                request=request
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Kullanıcı "{user.username}" {durum} yapıldı.'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Geçersiz istek'})


@login_required
def change_user_role(request, pk):
    """
    Kullanıcı rolünü değiştir
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Yetki hatası'})
    
    if request.method == 'POST':
        try:
            user = get_object_or_404(User, pk=pk)
            
            # Kendisinin rolünü değiştirmesini engelle
            if user == request.user:
                return JsonResponse({'success': False, 'error': 'Kendi rolünüzü değiştiremezsiniz'})
            
            import json
            data = json.loads(request.body)
            role = data.get('role', 'viewer')
            
            # Eski değerleri kaydet
            eski_yetkiler = {
                'is_superuser': user.is_superuser,
                'is_staff': user.is_staff,
                'role': getattr(user.profile, 'role', 'viewer') if hasattr(user, 'profile') else 'viewer'
            }
            
            # Role ayarla
            if role == 'super_user':
                user.is_superuser = True
                user.is_staff = True
                role_name = 'Süper Kullanıcı'
            elif role == 'province_admin':
                user.is_staff = True
                user.is_superuser = False
                role_name = 'İl Yöneticisi'
            elif role == 'province_manager':
                user.is_staff = True
                user.is_superuser = False
                role_name = 'İl Sorumlusu'
            else:  # viewer
                user.is_staff = False
                user.is_superuser = False
                role_name = 'Seyirci'
                
            user.save()
            
            # UserProfile güncelle
            profile, created = UserProfile.objects.get_or_create(user=user)
            profile.role = role
            profile.save()
            
            # Log kaydet
            log_user_activity(
                user=user,
                islem_yapan=request.user,
                islem_tipi='kullanici_rol_degisti',
                aciklama=f'Kullanıcı rolü {role_name} yapıldı: {user.username}',
                request=request,
                eski_deger=eski_yetkiler,
                yeni_deger={'is_superuser': user.is_superuser, 'is_staff': user.is_staff, 'role': role}
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Kullanıcı "{user.username}" rolü {role_name} olarak güncellendi.'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Geçersiz istek'})


@login_required
def get_user_cities(request, pk):
    """
    Kullanıcının sorumlu olduğu illeri getir
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Yetki hatası'})
    
    try:
        user = get_object_or_404(User, pk=pk)
        profile, created = UserProfile.objects.get_or_create(user=user)
        
        cities = profile.get_sorumlu_iller_list()
        
        return JsonResponse({
            'success': True,
            'cities': cities
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def assign_cities_to_user(request, pk):
    """
    Kullanıcıya il atama yap
    """
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Yetki hatası'})
    
    if request.method == 'POST':
        try:
            user = get_object_or_404(User, pk=pk)
            profile, created = UserProfile.objects.get_or_create(user=user)
            
            import json
            data = json.loads(request.body)
            cities = data.get('cities', [])
            
            # İl atamasını yap
            profile.set_sorumlu_iller(cities)
            profile.save()
            
            # Log kaydet
            if cities:
                cities_str = ', '.join(cities)
                log_message = f'Kullanıcıya il ataması yapıldı: {user.username} - İller: {cities_str}'
            else:
                log_message = f'Kullanıcının il ataması kaldırıldı: {user.username} (Tüm illere erişim)'
            
            log_user_activity(
                user=user,
                islem_yapan=request.user,
                islem_tipi='kullanici_il_atamasi',
                aciklama=log_message,
                request=request
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Kullanıcı "{user.username}" için il ataması güncellendi.'
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Geçersiz istek'})


@login_required
def generate_fake_tc(request):
    """
    Test amaçlı sahte TC kimlik numarası oluştur
    """
    import random
    
    # İlk 9 hanesi rastgele
    first_nine = [random.randint(1, 9)] + [random.randint(0, 9) for _ in range(8)]
    
    # 10. hane hesaplanması
    odd_sum = sum(first_nine[i] for i in range(0, 9, 2))  # 1,3,5,7,9. haneler
    even_sum = sum(first_nine[i] for i in range(1, 8, 2))  # 2,4,6,8. haneler
    
    tenth_digit = ((odd_sum * 7) - even_sum) % 10
    
    # 11. hane hesaplanması
    first_ten_sum = sum(first_nine) + tenth_digit
    eleventh_digit = first_ten_sum % 10
    
    # Tam TC kimlik numarası
    tc_kimlik = ''.join(map(str, first_nine)) + str(tenth_digit) + str(eleventh_digit)
    
    return JsonResponse({'tc_kimlik': tc_kimlik})
